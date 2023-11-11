import inspect
import logging
import softest
import csv
from openpyxl import Workbook, load_workbook

import logging

# create logger instance
logger_instance = logging.getLogger(__name__)
logger_instance.setLevel(level=logging.INFO)
formatter = logging.Formatter('[%(levelname)s] %(asctime)s %(module)s(line: %(lineno)d): %(message)s'
                              , "%Y-%m-%d %H:%M:%S")

# create streamHandler instance
streamer = logging.StreamHandler()
streamer.setLevel(level=logging.INFO)
streamer.setFormatter(formatter)
logger_instance.addHandler(streamer)


class Utils(softest.TestCase):
    # def assertListItemText(self, list_item, value):
    #     for stop in list_item:
    #         self.soft_assert(self.assertEqual, stop.text, value)
    #         if stop.text == value:
    #             print("result = test passed")
    #         else:
    #             print("result = test failed")
    #     self.assert_all()

    def custom_logger(logLevel=logging.DEBUG):
        logger = logging.getLogger(__name__)
        logger.setLevel(level=logLevel)
        formatter = logging.Formatter('[%(levelname)s] %(asctime)s %(module)s(line: %(lineno)d): %(message)s', "%m/%d/%Y %H:%M:%S")
        streamer = logging.StreamHandler()
        streamer.setLevel(level=logLevel)
        streamer.setFormatter(formatter)
        logger.addHandler(streamer)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        #Create an empty list
        datalist = []
        #Open CSV file
        csvdata = open(filename, "r")
        #Create CSV reader
        reader = csv.reader(csvdata)
        #skip header
        next(reader)
        #Add CSV rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist
